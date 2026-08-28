from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Iterable, Optional

from sqlalchemy import func, select
from sqlalchemy.orm import Session, joinedload

from app.engines.fx import translate_amount
from app.config import get_settings
from app.engines.entity_close import is_journal_led_entity
from app.engines.intercompany import account_is_ic_leg
from app.models import BankAccount, DimAccount, DimEntity, DimReportLayout, DimScenario, Transaction
from app.schemas.reports import (
    AnalyticsKpi,
    AnalyticsPack,
    FluxItem,
    ReportFilter,
    ReportLine,
    ReportOut,
)

FYE_MONTH = 7  # Fiscal year-end 31 July
BALANCE_TOLERANCE = Decimal("0.02")
NIL_TOLERANCE = Decimal("0.005")
CURRENT_EARNINGS_CODE = "BS_CURRENT_EARNINGS"
TRADE_AR_CODE = "BS_AR"
TRADE_AP_CODE = "BS_AP"
IC_LINE_CODE = "BS_IC"
CASH_LINE_CODE = "BS_CASH"
CASH_XFER_CODE = "BS_CASH_XFER"
EQUITY_CODE = "BS_EQUITY"
PNL_TITLE = "Profit & Loss"
BS_TITLE = "Balance Sheet"

# Cashbook BS: liabilities keep the bank sign; non-cash assets flip so purchases increase the asset.
CASHBOOK_LIABILITY_LINES = {
    TRADE_AP_CODE,
    IC_LINE_CODE,
    "BS_UNEARNED",
    "BS_TAX",
    "BS_SH_LOAN",
    CASH_XFER_CODE,
}
CASHBOOK_ASSET_FLIP_LINES = {
    TRADE_AR_CODE,
    "BS_INV",
    "BS_PREPAID",
    "BS_FA",
}


def _month_end(year: int, month: int) -> date:
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def _add_months(year: int, month: int, delta: int) -> tuple[int, int]:
    idx = year * 12 + (month - 1) + delta
    return idx // 12, idx % 12 + 1


def _safe_replace_year(d: date, year: int) -> date:
    try:
        return d.replace(year=year)
    except ValueError:
        return date(year, d.month, 28)


def _pretty_date(d: date) -> str:
    return f"{d.day} {d.strftime('%B %Y')}"


def fiscal_year_start(year: int, month: int) -> date:
    """First day of the fiscal year that contains year-month (FYE 31 July)."""
    if month <= FYE_MONTH:
        return date(year - 1, FYE_MONTH + 1, 1)
    return date(year, FYE_MONTH + 1, 1)


def period_label(filters: ReportFilter) -> str:
    year = filters.year or (filters.as_of_date.year if filters.as_of_date else date.today().year)
    month = filters.month or (filters.as_of_date.month if filters.as_of_date else date.today().month)
    end = filters.as_of_date or filters.date_to or _month_end(year, month)
    if filters.report_type == "balance_sheet":
        return f"As at {_pretty_date(end)}"
    if filters.report_type == "equity":
        return f"Fiscal YTD ended {_pretty_date(end)}"
    if filters.period == "monthly":
        return f"Month ended {_pretty_date(_month_end(year, month))}"
    if filters.period == "quarterly":
        q = filters.quarter or ((month - 1) // 3 + 1)
        return f"Quarter ended {_pretty_date(_month_end(year, q * 3))}"
    if filters.period == "ytd":
        return f"Fiscal YTD ended {_pretty_date(end)}"
    start = filters.date_from or fiscal_year_start(year, month)
    return f"{_pretty_date(start)} – {_pretty_date(end)}"


def _as_of_date(filters: ReportFilter) -> date:
    if filters.as_of_date:
        return filters.as_of_date
    if filters.date_to:
        return filters.date_to
    year = filters.year or date.today().year
    month = filters.month or date.today().month
    return _month_end(year, month)


def use_cashbook_presentation(db: Session, filters: ReportFilter) -> bool:
    """Cashbook BS only when every selected entity is synoptic-led (not USA journals)."""
    if filters.report_type != "balance_sheet":
        return False
    entity_ids = filters.entity_ids or []
    if not entity_ids:
        return False
    return all(not is_journal_led_entity(db, eid) for eid in entity_ids)


def _entity_banks(db: Session, entity_ids: list[int] | None) -> list[BankAccount]:
    q = select(BankAccount).where(BankAccount.is_active.is_(True))
    if entity_ids:
        q = q.where(BankAccount.entity_id.in_(entity_ids))
    return list(db.scalars(q).all())


def _bank_book_balance(db: Session, bank: BankAccount, as_of: date, scenario_id: int) -> Decimal:
    total = db.scalar(
        select(func.coalesce(func.sum(Transaction.amount), 0)).where(
            Transaction.bank_account_id == bank.id,
            Transaction.txn_date <= as_of,
            Transaction.scenario_id == scenario_id,
            Transaction.status.notin_(["void", "excluded"]),
        )
    )
    return Decimal(bank.opening_balance) + Decimal(total or 0)


def _translate_closing(
    db: Session,
    *,
    amount: Decimal,
    from_currency: str,
    to_currency: str,
    as_of: date,
) -> tuple[Decimal, bool, str | None]:
    translated = translate_amount(
        db,
        amount=amount,
        from_currency=from_currency,
        to_currency=to_currency,
        as_of=as_of,
        rate_type="closing",
    )
    pair = None
    if translated.missing:
        pair = f"{from_currency}→{to_currency}"
    return translated.amount, translated.missing, pair


def cashbook_book_cash(db: Session, filters: ReportFilter) -> tuple[Decimal, bool, list[str]]:
    """Reporting-currency sum of bank books (opening + activity) at closing rate."""
    as_of = _as_of_date(filters)
    reporting = filters.reporting_currency or get_settings().default_reporting_currency
    total = Decimal("0")
    missing = False
    pairs: list[str] = []
    seen: set[str] = set()
    for bank in _entity_banks(db, filters.entity_ids):
        native = _bank_book_balance(db, bank, as_of, filters.scenario_id)
        amount, miss, pair = _translate_closing(
            db,
            amount=native,
            from_currency=bank.currency,
            to_currency=reporting,
            as_of=as_of,
        )
        total += amount
        if miss and pair and pair not in seen:
            missing = True
            seen.add(pair)
            pairs.append(pair)
    return total, missing, pairs


def _cashbook_opening_cash(db: Session, filters: ReportFilter) -> tuple[Decimal, bool, list[str]]:
    as_of = _as_of_date(filters)
    reporting = filters.reporting_currency or get_settings().default_reporting_currency
    total = Decimal("0")
    missing = False
    pairs: list[str] = []
    seen: set[str] = set()
    for bank in _entity_banks(db, filters.entity_ids):
        amount, miss, pair = _translate_closing(
            db,
            amount=Decimal(bank.opening_balance),
            from_currency=bank.currency,
            to_currency=reporting,
            as_of=as_of,
        )
        total += amount
        if miss and pair and pair not in seen:
            missing = True
            seen.add(pair)
            pairs.append(pair)
    return total, missing, pairs


def _cashbook_signed_amount(account: DimAccount, amount: Decimal, line_code: str) -> Decimal:
    if line_code in CASHBOOK_LIABILITY_LINES:
        return amount
    if line_code in CASHBOOK_ASSET_FLIP_LINES:
        return -amount
    return _signed_amount(account, amount, "balance_sheet")


def _pnl_from_totals(
    db: Session,
    totals: dict[int, Decimal],
    accounts: dict[int, DimAccount],
) -> Decimal:
    layouts = list(
        db.scalars(
            select(DimReportLayout)
            .where(DimReportLayout.report_type == "income_statement")
            .order_by(DimReportLayout.sort_order.asc())
        )
    )
    if not layouts:
        rev = Decimal("0")
        exp = Decimal("0")
        for acct in accounts.values():
            raw = totals.get(acct.id, Decimal("0"))
            if acct.account_type == "revenue":
                rev += _signed_amount(acct, raw, "income_statement")
            elif acct.account_type == "expense":
                exp += _signed_amount(acct, raw, "income_statement")
        return rev - exp
    line_amounts: dict[str, Decimal] = {}
    for layout in layouts:
        amount, _, _ = _compute_layout_amount(layout, totals, accounts, "income_statement", line_amounts)
        line_amounts[layout.line_code] = amount
    for code in ("NI", "NET_INCOME"):
        if code in line_amounts:
            return line_amounts[code]
    return line_amounts.get("TOT_REV", Decimal("0")) - line_amounts.get("TOT_EXP", Decimal("0"))


def _pnl_between(
    db: Session,
    filters: ReportFilter,
    accounts: dict[int, DimAccount],
    date_from: date,
    date_to: date,
) -> Decimal:
    if date_from > date_to:
        return Decimal("0")
    scoped = filters.model_copy(
        update={
            "report_type": "income_statement",
            "period": "custom",
            "date_from": date_from,
            "date_to": date_to,
            "as_of_date": date_to,
            "year": date_to.year,
            "month": date_to.month,
            "compare_prior_period": False,
            "compare_prior_year": False,
            "compare_budget": False,
            "compare_scenario_id": None,
        }
    )
    totals = aggregate_by_account(db, scoped, balance_sheet=False)
    return _pnl_from_totals(db, totals, accounts)


def cashbook_bs_overrides(
    db: Session,
    filters: ReportFilter,
    accounts: dict[int, DimAccount],
    totals: dict[int, Decimal],
) -> tuple[dict[str, Decimal], bool, list[str]]:
    """Book cash, GL 1000 as due-to-other-banks, opening equity (openings + RE + pre-FY NI)."""
    cash, miss_c, pairs_c = cashbook_book_cash(db, filters)
    opening, miss_o, pairs_o = _cashbook_opening_cash(db, filters)
    missing = miss_c or miss_o
    pairs = list(pairs_c)
    for p in pairs_o:
        if p not in pairs:
            pairs.append(p)

    acct_1000 = next((a for a in accounts.values() if a.code == "1000"), None)
    xfer = totals.get(acct_1000.id, Decimal("0")) if acct_1000 else Decimal("0")

    acct_3000 = next((a for a in accounts.values() if a.code == "3000"), None)
    gl_re = Decimal("0")
    if acct_3000:
        gl_re = _signed_amount(acct_3000, totals.get(acct_3000.id, Decimal("0")), "balance_sheet")

    as_of = _as_of_date(filters)
    year = filters.year or as_of.year
    month = filters.month or as_of.month
    fy_start = fiscal_year_start(year, month)
    pre_fy = _pnl_between(db, filters, accounts, date(2000, 1, 1), fy_start - timedelta(days=1))

    return (
        {
            CASH_LINE_CODE: cash,
            CASH_XFER_CODE: xfer,
            EQUITY_CODE: opening + gl_re + pre_fy,
        },
        missing,
        pairs,
    )


def _is_section_header(line: ReportLine) -> bool:
    return (not line.is_total) and (not line.drillable) and line.indent_level == 0


def _line_has_amount(line: ReportLine) -> bool:
    for val in (
        line.amount,
        line.prior_period_amount,
        line.prior_year_amount,
        line.budget_amount,
        line.compare_amount,
    ):
        if val is not None and abs(val) > NIL_TOLERANCE:
            return True
    return False


def prune_zero_report_lines(lines: list[ReportLine]) -> list[ReportLine]:
    """Drop nil leaf rows; keep totals and section headers that still have a body."""
    kept: list[ReportLine] = []
    for line in lines:
        if line.is_total or _is_section_header(line) or _line_has_amount(line):
            kept.append(line)
    out: list[ReportLine] = []
    i = 0
    while i < len(kept):
        line = kept[i]
        if _is_section_header(line):
            j = i + 1
            has_body = False
            while j < len(kept) and not _is_section_header(kept[j]):
                child = kept[j]
                if child.is_total or _line_has_amount(child):
                    has_body = True
                    break
                j += 1
            if has_body:
                out.append(line)
        else:
            out.append(line)
        i += 1
    return out


def prior_period_filters(filters: ReportFilter) -> ReportFilter:
    year = filters.year or (filters.as_of_date.year if filters.as_of_date else date.today().year)
    month = filters.month or (filters.as_of_date.month if filters.as_of_date else date.today().month)
    if filters.period == "quarterly" and filters.report_type != "balance_sheet":
        q = filters.quarter or ((month - 1) // 3 + 1)
        if q == 1:
            return filters.model_copy(update={"year": year - 1, "quarter": 4, "month": 12})
        end_month = q * 3
        start_month = end_month - 2
        prior_end = start_month - 1
        return filters.model_copy(update={"quarter": q - 1, "month": prior_end})
    py, pm = _add_months(year, month, -1)
    as_of = _month_end(py, pm)
    return filters.model_copy(
        update={
            "year": py,
            "month": pm,
            "as_of_date": as_of,
            "date_to": as_of,
            "date_from": None,
        }
    )


def prior_year_filters(filters: ReportFilter) -> ReportFilter:
    year = (filters.year or (filters.as_of_date.year if filters.as_of_date else date.today().year)) - 1
    as_of = _safe_replace_year(filters.as_of_date, filters.as_of_date.year - 1) if filters.as_of_date else None
    date_to = _safe_replace_year(filters.date_to, filters.date_to.year - 1) if filters.date_to else as_of
    date_from = _safe_replace_year(filters.date_from, filters.date_from.year - 1) if filters.date_from else None
    return filters.model_copy(
        update={
            "year": year,
            "as_of_date": as_of,
            "date_to": date_to,
            "date_from": date_from,
        }
    )


def _variance_pct(amount: Decimal, compare: Decimal | None) -> Decimal | None:
    if compare is None or compare == 0:
        return None
    return (amount - compare) / abs(compare) * Decimal("100")


def _budget_scenario_id(db: Session) -> int | None:
    row = db.scalar(select(DimScenario).where(DimScenario.code == "BUDGET"))
    return row.id if row else None


def _period_bounds(filters: ReportFilter) -> tuple[date, date]:
    today = date.today()
    year = filters.year or today.year

    if filters.period == "monthly":
        month = filters.month or today.month
        return date(year, month, 1), _month_end(year, month)

    if filters.period == "quarterly":
        q = filters.quarter or ((today.month - 1) // 3 + 1)
        start_month = (q - 1) * 3 + 1
        end_month = start_month + 2
        return date(year, start_month, 1), _month_end(year, end_month)

    if filters.period == "ytd":
        month = filters.month or (filters.as_of_date.month if filters.as_of_date else today.month)
        year = filters.year or (filters.as_of_date.year if filters.as_of_date else today.year)
        end = filters.date_to or filters.as_of_date or _month_end(year, month)
        return fiscal_year_start(year, month), end

    # custom
    month = filters.month or today.month
    start = filters.date_from or fiscal_year_start(year, month)
    end = filters.date_to or filters.as_of_date or today
    return start, end


def _iter_fact_lines(
    db: Session,
    *,
    date_from: date,
    date_to: date,
    scenario_id: int,
    entity_ids: Optional[list[int]],
    department_ids: Optional[list[int]],
) -> Iterable[tuple[Transaction, int, Optional[int], Decimal]]:
    """
    Yield (source_txn, account_id, department_id, amount) fact grains.
    Split transactions explode into multiple fact lines.
    """
    q = (
        select(Transaction)
        .options(
            joinedload(Transaction.splits),
            joinedload(Transaction.bank_account),
            joinedload(Transaction.entity),
        )
        .where(
            Transaction.scenario_id == scenario_id,
            Transaction.txn_date >= date_from,
            Transaction.txn_date <= date_to,
            Transaction.status.notin_(["void", "excluded"]),
        )
    )
    if entity_ids:
        q = q.where(Transaction.entity_id.in_(entity_ids))
    if department_ids:
        # Filter parent dept OR any split dept — applied below for splits
        pass

    txns = db.scalars(q).unique().all()
    for txn in txns:
        if txn.is_split and txn.splits:
            for split in txn.splits:
                if department_ids and split.department_id not in department_ids:
                    continue
                yield txn, split.account_id, split.department_id, Decimal(split.amount)
        else:
            if txn.account_id is None:
                continue
            if department_ids and txn.department_id not in department_ids:
                continue
            yield txn, txn.account_id, txn.department_id, Decimal(txn.amount)


def _fact_memo(txn: Transaction, account_id: int, amount: Decimal) -> str:
    if txn.is_split and txn.splits:
        for split in txn.splits:
            if split.account_id == account_id:
                return split.memo or ""
    return txn.memo or ""


@dataclass
class AccountBundle:
    totals: dict[int, Decimal]
    ic_totals: dict[int, Decimal]
    fx_missing: bool = False
    fx_missing_pairs: list[str] = field(default_factory=list)


def aggregate_account_bundle(
    db: Session,
    filters: ReportFilter,
    *,
    balance_sheet: bool = False,
) -> AccountBundle:
    """Account totals in reporting currency, with IC legs split out of 1100/2000."""
    settings = get_settings()
    reporting_currency = filters.reporting_currency or settings.default_reporting_currency
    accounts = {a.id: a for a in db.scalars(select(DimAccount)).all()}

    if balance_sheet:
        date_from = date(2000, 1, 1)
        date_to = filters.as_of_date or filters.date_to or date.today()
    else:
        date_from, date_to = _period_bounds(filters)

    totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    ic_totals: dict[int, Decimal] = defaultdict(lambda: Decimal("0"))
    missing_pairs: list[str] = []
    seen_missing: set[str] = set()

    for txn, account_id, _dept, amount in _iter_fact_lines(
        db,
        date_from=date_from,
        date_to=date_to,
        scenario_id=filters.scenario_id,
        entity_ids=filters.entity_ids,
        department_ids=filters.department_ids,
    ):
        translated = translate_amount(
            db,
            amount=amount,
            from_currency=txn.currency,
            to_currency=reporting_currency,
            as_of=txn.txn_date,
            rate_type="average" if not balance_sheet else "closing",
        )
        if translated.missing:
            pair = f"{txn.currency}→{reporting_currency}"
            if pair not in seen_missing:
                seen_missing.add(pair)
                missing_pairs.append(pair)
        totals[account_id] += translated.amount
        acct = accounts.get(account_id)
        extra = _fact_memo(txn, account_id, amount)
        if account_is_ic_leg(acct, txn, extra):
            ic_totals[account_id] += translated.amount

    return AccountBundle(
        totals=dict(totals),
        ic_totals=dict(ic_totals),
        fx_missing=bool(missing_pairs),
        fx_missing_pairs=missing_pairs,
    )


def aggregate_by_account(
    db: Session,
    filters: ReportFilter,
    *,
    balance_sheet: bool = False,
) -> dict[int, Decimal]:
    """Return {account_id: amount} in reporting currency."""
    return aggregate_account_bundle(db, filters, balance_sheet=balance_sheet).totals


def _signed_amount(account: DimAccount, amount: Decimal, report_type: str) -> Decimal:
    """Normalize bank-signed amounts into statement presentation."""
    # Bank convention: +inflow / -outflow
    # For P&L: revenue inflows positive, expenses outflows as positive expense
    if report_type == "income_statement":
        if account.account_type == "revenue":
            return amount  # inflows already positive
        if account.account_type == "expense":
            return -amount  # outflows negative → show as positive expense
        return amount
    if report_type == "balance_sheet":
        if account.normal_balance == "credit":
            return -amount
        return amount
    return amount


def _pnl_account_ids(accounts: dict[int, DimAccount]) -> list[int]:
    return [a.id for a in accounts.values() if a.account_type in ("revenue", "expense") and a.is_active]


def _compute_layout_amount(
    layout: DimReportLayout,
    totals: dict[int, Decimal],
    accounts: dict[int, DimAccount],
    report_type: str,
    line_amounts: dict[str, Decimal],
    *,
    ic_totals: dict[int, Decimal] | None = None,
    overrides: dict[str, Decimal] | None = None,
    cashbook: bool = False,
) -> tuple[Decimal, list[int], str | None]:
    drill_ids: list[int] = []
    type_filter = layout.account_type_filter
    amount = Decimal("0")
    ic_totals = ic_totals or {}
    accounts_by_code = {a.code: a for a in accounts.values()}

    def _present(acct: DimAccount | None, raw: Decimal, line_code: str) -> Decimal:
        if acct is None:
            return -raw if cashbook and line_code in CASHBOOK_ASSET_FLIP_LINES else raw
        if cashbook and report_type == "balance_sheet" and line_code != IC_LINE_CODE:
            return _cashbook_signed_amount(acct, raw, line_code)
        return _signed_amount(acct, raw, report_type)

    if overrides and layout.line_code in overrides:
        amount = overrides[layout.line_code]
        if layout.line_code == CURRENT_EARNINGS_CODE:
            drill_ids = _pnl_account_ids(accounts)
        elif layout.account_id:
            drill_ids = [layout.account_id]
        elif layout.line_code == CASH_XFER_CODE:
            acct = accounts_by_code.get("1000")
            if acct:
                drill_ids = [acct.id]
        return amount, drill_ids, type_filter

    if layout.line_code == TRADE_AR_CODE and layout.account_id:
        acct = accounts.get(layout.account_id)
        raw = totals.get(layout.account_id, Decimal("0")) - ic_totals.get(layout.account_id, Decimal("0"))
        amount = _present(acct, raw, layout.line_code)
        drill_ids = [layout.account_id]
        return amount, drill_ids, type_filter

    if layout.line_code == TRADE_AP_CODE and layout.account_id:
        acct = accounts.get(layout.account_id)
        raw = totals.get(layout.account_id, Decimal("0")) - ic_totals.get(layout.account_id, Decimal("0"))
        amount = _present(acct, raw, layout.line_code)
        drill_ids = [layout.account_id]
        return amount, drill_ids, type_filter

    if layout.line_code == IC_LINE_CODE:
        ic_acct = accounts.get(layout.account_id) if layout.account_id else accounts_by_code.get("2100")
        raw_2100 = totals.get(ic_acct.id, Decimal("0")) if ic_acct else Decimal("0")
        amount = _signed_amount(ic_acct, raw_2100, report_type) if ic_acct else raw_2100
        ar = accounts_by_code.get("1100")
        ap = accounts_by_code.get("2000")
        if ar:
            # Due-from (asset) nets against due-to on the IC liability line.
            amount -= _signed_amount(ar, ic_totals.get(ar.id, Decimal("0")), report_type)
            drill_ids.append(ar.id)
        if ap:
            amount += _signed_amount(ap, ic_totals.get(ap.id, Decimal("0")), report_type)
            drill_ids.append(ap.id)
        if ic_acct:
            drill_ids.append(ic_acct.id)
        return amount, drill_ids, type_filter

    if layout.account_id:
        raw = totals.get(layout.account_id, Decimal("0"))
        acct = accounts.get(layout.account_id)
        amount = _present(acct, raw, layout.line_code)
        if layout.sign_flip:
            amount = -amount
        drill_ids = [layout.account_id]
    elif layout.account_type_filter:
        for acct_id, raw in totals.items():
            acct = accounts.get(acct_id)
            if acct and acct.account_type == layout.account_type_filter:
                amount += _present(acct, raw, layout.line_code)
                drill_ids.append(acct_id)
        if not drill_ids:
            drill_ids = [a.id for a in accounts.values() if a.account_type == layout.account_type_filter]
    elif layout.calc_formula:
        amount = _eval_formula(layout.calc_formula, line_amounts)
        if layout.line_code in ("NI", "NET_INCOME") or "NET" in (layout.line_label or "").upper():
            drill_ids = _pnl_account_ids(accounts)
        elif layout.section in ("revenue", "expense", "asset", "liability", "equity"):
            drill_ids = [a.id for a in accounts.values() if a.account_type == layout.section and a.is_active]
            type_filter = layout.section

    return amount, drill_ids, type_filter


def _series_amount(
    layout: DimReportLayout,
    totals: dict[int, Decimal] | None,
    accounts: dict[int, DimAccount],
    report_type: str,
    line_amounts: dict[str, Decimal],
    *,
    ic_totals: dict[int, Decimal] | None = None,
    overrides: dict[str, Decimal] | None = None,
    cashbook: bool = False,
) -> Decimal | None:
    if totals is None:
        return None
    amount, _, _ = _compute_layout_amount(
        layout,
        totals,
        accounts,
        report_type,
        line_amounts,
        ic_totals=ic_totals,
        overrides=overrides,
        cashbook=cashbook,
    )
    line_amounts[layout.line_code] = amount
    return amount


def _materiality(filters: ReportFilter) -> tuple[Decimal, Decimal]:
    amt = filters.materiality_amount if filters.materiality_amount is not None else Decimal("1000")
    pct = filters.materiality_pct if filters.materiality_pct is not None else Decimal("10")
    return amt, pct


def _flag_flux(
    amount: Decimal,
    prior: Decimal | None,
    *,
    mat_amt: Decimal,
    mat_pct: Decimal,
) -> str | None:
    if prior is None:
        return None
    if abs(amount) < Decimal("0.005") and abs(prior) < Decimal("0.005"):
        return None
    if abs(prior) < Decimal("0.005") and abs(amount) >= mat_amt:
        return "new"
    if abs(amount) < Decimal("0.005") and abs(prior) >= mat_amt:
        return "drop"
    variance = amount - prior
    pct = _variance_pct(amount, prior)
    if abs(variance) >= mat_amt or (pct is not None and abs(pct) >= mat_pct):
        return "material"
    return None


def _flux_note(label: str, amount: Decimal, prior: Decimal | None, flag: str | None) -> str | None:
    if not flag:
        return None
    if flag == "new":
        return f"{label} is new this period ({amount:,.2f})."
    if flag == "drop":
        return f"{label} dropped to nil (was {prior:,.2f})."
    variance = amount - (prior or Decimal("0"))
    pct = _variance_pct(amount, prior)
    direction = "up" if variance > 0 else "down"
    pct_txt = f" ({pct:.1f}%)" if pct is not None else ""
    return f"{label} {direction} {abs(variance):,.2f}{pct_txt} vs prior period."


def _flux_items(report: ReportOut) -> list[FluxItem]:
    items: list[FluxItem] = []
    for line in report.lines:
        if not line.flux_flag:
            continue
        if line.indent_level == 0 and not line.is_total and not line.drillable:
            continue
        items.append(
            FluxItem(
                report_type=report.report_type,
                line_code=line.line_code,
                line_label=line.line_label,
                wp_ref=line.wp_ref,
                amount=line.amount,
                prior_amount=line.prior_period_amount,
                variance=line.prior_period_variance,
                variance_pct=line.prior_period_variance_pct,
                flag=line.flux_flag,
                note=line.flux_note or "",
                drillable=line.drillable,
            )
        )
    items.sort(key=lambda i: abs(i.variance or Decimal("0")), reverse=True)
    return items


def _entity_cover_name(db: Session, filters: ReportFilter) -> str:
    if not filters.entity_ids:
        return "All entities" if filters.consolidate else "Entity"
    rows = list(db.scalars(select(DimEntity).where(DimEntity.id.in_(filters.entity_ids))).all())
    if not rows:
        return "Entity"
    if len(rows) == 1:
        return rows[0].name or rows[0].code
    return " + ".join(e.code for e in rows)


def _statement_title(report_type: str) -> str:
    return {
        "income_statement": PNL_TITLE,
        "balance_sheet": BS_TITLE,
        "equity": "Statement of Changes in Equity",
    }.get(report_type, report_type)


def _ytd_filters(filters: ReportFilter) -> ReportFilter:
    year = filters.year or (filters.as_of_date.year if filters.as_of_date else date.today().year)
    month = filters.month or (filters.as_of_date.month if filters.as_of_date else date.today().month)
    end = filters.as_of_date or filters.date_to or _month_end(year, month)
    start = fiscal_year_start(year, month)
    return filters.model_copy(
        update={
            "report_type": "income_statement",
            "period": "ytd",
            "year": year,
            "month": month,
            "date_from": start,
            "date_to": end,
            "as_of_date": end,
            "compare_prior_period": False,
            "compare_prior_year": False,
            "compare_budget": False,
            "compare_scenario_id": None,
        }
    )


def _pnl_net_income(
    db: Session,
    filters: ReportFilter,
    accounts: dict[int, DimAccount],
    *,
    totals: dict[int, Decimal] | None = None,
) -> Decimal:
    ytd = _ytd_filters(filters)
    if totals is None:
        totals = aggregate_by_account(db, ytd, balance_sheet=False)
    return _pnl_from_totals(db, totals, accounts)


def build_report(db: Session, filters: ReportFilter) -> ReportOut:
    report_type = filters.report_type
    balance_sheet = report_type == "balance_sheet"
    bundle = aggregate_account_bundle(db, filters, balance_sheet=balance_sheet)
    totals = bundle.totals
    ic_totals = bundle.ic_totals
    fx_missing = bundle.fx_missing
    fx_missing_pairs = list(bundle.fx_missing_pairs)

    compare_bundle: AccountBundle | None = None
    if filters.compare_scenario_id:
        compare_filters = filters.model_copy(update={"scenario_id": filters.compare_scenario_id})
        compare_bundle = aggregate_account_bundle(db, compare_filters, balance_sheet=balance_sheet)
        fx_missing = fx_missing or compare_bundle.fx_missing
        fx_missing_pairs += [p for p in compare_bundle.fx_missing_pairs if p not in fx_missing_pairs]
    compare_totals = compare_bundle.totals if compare_bundle else None
    compare_ic = compare_bundle.ic_totals if compare_bundle else None

    prior_bundle: AccountBundle | None = None
    prior_filters = None
    if filters.compare_prior_period:
        prior_filters = prior_period_filters(filters)
        prior_bundle = aggregate_account_bundle(db, prior_filters, balance_sheet=balance_sheet)
        fx_missing = fx_missing or prior_bundle.fx_missing
        fx_missing_pairs += [p for p in prior_bundle.fx_missing_pairs if p not in fx_missing_pairs]
    prior_totals = prior_bundle.totals if prior_bundle else None
    prior_ic = prior_bundle.ic_totals if prior_bundle else None

    py_bundle: AccountBundle | None = None
    py_filters = None
    if filters.compare_prior_year:
        py_filters = prior_year_filters(filters)
        py_bundle = aggregate_account_bundle(db, py_filters, balance_sheet=balance_sheet)
        fx_missing = fx_missing or py_bundle.fx_missing
        fx_missing_pairs += [p for p in py_bundle.fx_missing_pairs if p not in fx_missing_pairs]
    py_totals = py_bundle.totals if py_bundle else None
    py_ic = py_bundle.ic_totals if py_bundle else None

    budget_bundle: AccountBundle | None = None
    budget_id = _budget_scenario_id(db) if filters.compare_budget else None
    if budget_id:
        budget_filters = filters.model_copy(update={"scenario_id": budget_id})
        budget_bundle = aggregate_account_bundle(db, budget_filters, balance_sheet=balance_sheet)
        fx_missing = fx_missing or budget_bundle.fx_missing
        fx_missing_pairs += [p for p in budget_bundle.fx_missing_pairs if p not in fx_missing_pairs]
    budget_totals = budget_bundle.totals if budget_bundle else None
    budget_ic = budget_bundle.ic_totals if budget_bundle else None

    accounts = {a.id: a for a in db.scalars(select(DimAccount)).all()}
    layouts = list(
        db.scalars(
            select(DimReportLayout)
            .where(DimReportLayout.report_type == report_type)
            .order_by(DimReportLayout.sort_order.asc())
        )
    )

    earnings_overrides: dict[str, Decimal] = {}
    compare_overrides: dict[str, Decimal] = {}
    prior_overrides: dict[str, Decimal] = {}
    py_overrides: dict[str, Decimal] = {}
    budget_overrides: dict[str, Decimal] = {}
    cashbook = use_cashbook_presentation(db, filters) if balance_sheet else False
    if balance_sheet:
        earnings_overrides[CURRENT_EARNINGS_CODE] = _pnl_net_income(db, filters, accounts)
        earnings_overrides[CASH_XFER_CODE] = Decimal("0")
        if cashbook:
            cb, miss, pairs = cashbook_bs_overrides(db, filters, accounts, totals)
            earnings_overrides.update(cb)
            fx_missing = fx_missing or miss
            fx_missing_pairs += [p for p in pairs if p not in fx_missing_pairs]
        if compare_totals is not None:
            compare_filters = filters.model_copy(update={"scenario_id": filters.compare_scenario_id})
            compare_overrides[CURRENT_EARNINGS_CODE] = _pnl_net_income(db, compare_filters, accounts)
            compare_overrides[CASH_XFER_CODE] = Decimal("0")
            if cashbook:
                cb, miss, pairs = cashbook_bs_overrides(db, compare_filters, accounts, compare_totals)
                compare_overrides.update(cb)
                fx_missing = fx_missing or miss
                fx_missing_pairs += [p for p in pairs if p not in fx_missing_pairs]
        if prior_filters is not None and prior_totals is not None:
            prior_overrides[CURRENT_EARNINGS_CODE] = _pnl_net_income(db, prior_filters, accounts)
            prior_overrides[CASH_XFER_CODE] = Decimal("0")
            if cashbook:
                cb, miss, pairs = cashbook_bs_overrides(db, prior_filters, accounts, prior_totals)
                prior_overrides.update(cb)
                fx_missing = fx_missing or miss
                fx_missing_pairs += [p for p in pairs if p not in fx_missing_pairs]
        if py_filters is not None and py_totals is not None:
            py_overrides[CURRENT_EARNINGS_CODE] = _pnl_net_income(db, py_filters, accounts)
            py_overrides[CASH_XFER_CODE] = Decimal("0")
            if cashbook:
                cb, miss, pairs = cashbook_bs_overrides(db, py_filters, accounts, py_totals)
                py_overrides.update(cb)
                fx_missing = fx_missing or miss
                fx_missing_pairs += [p for p in pairs if p not in fx_missing_pairs]
        if budget_id and budget_totals is not None:
            budget_filters = filters.model_copy(update={"scenario_id": budget_id})
            budget_overrides[CURRENT_EARNINGS_CODE] = _pnl_net_income(db, budget_filters, accounts)
            budget_overrides[CASH_XFER_CODE] = Decimal("0")
            if cashbook:
                cb, miss, pairs = cashbook_bs_overrides(db, budget_filters, accounts, budget_totals)
                budget_overrides.update(cb)
                fx_missing = fx_missing or miss
                fx_missing_pairs += [p for p in pairs if p not in fx_missing_pairs]

    if not layouts:
        report = _synthesize_report(
            db,
            filters,
            totals,
            compare_totals or {},
            accounts,
            prior_totals=prior_totals,
            py_totals=py_totals,
            budget_totals=budget_totals,
            current_earnings=earnings_overrides.get(CURRENT_EARNINGS_CODE),
        )
        return _finalize_report(
            report,
            filters,
            prior_filters,
            py_filters,
            db=db,
            fx_missing=fx_missing,
            fx_missing_pairs=fx_missing_pairs,
        )

    line_amounts: dict[str, Decimal] = {}
    compare_line_amounts: dict[str, Decimal] = {}
    prior_line_amounts: dict[str, Decimal] = {}
    py_line_amounts: dict[str, Decimal] = {}
    budget_line_amounts: dict[str, Decimal] = {}
    lines: list[ReportLine] = []
    section_refs = {"revenue": "A", "expense": "B", "asset": "C", "liability": "D", "equity": "E", "totals": "Z"}
    section_counters: dict[str, int] = defaultdict(int)
    mat_amt, mat_pct = _materiality(filters)

    for layout in layouts:
        amount, drill_ids, type_filter = _compute_layout_amount(
            layout,
            totals,
            accounts,
            report_type,
            line_amounts,
            ic_totals=ic_totals,
            overrides=earnings_overrides,
            cashbook=cashbook,
        )
        line_amounts[layout.line_code] = amount

        compare_amount = _series_amount(
            layout,
            compare_totals,
            accounts,
            report_type,
            compare_line_amounts,
            ic_totals=compare_ic,
            overrides=compare_overrides or None,
            cashbook=cashbook,
        )
        prior_amount = _series_amount(
            layout,
            prior_totals,
            accounts,
            report_type,
            prior_line_amounts,
            ic_totals=prior_ic,
            overrides=prior_overrides or None,
            cashbook=cashbook,
        )
        py_amount = _series_amount(
            layout,
            py_totals,
            accounts,
            report_type,
            py_line_amounts,
            ic_totals=py_ic,
            overrides=py_overrides or None,
            cashbook=cashbook,
        )
        budget_amount = _series_amount(
            layout,
            budget_totals,
            accounts,
            report_type,
            budget_line_amounts,
            ic_totals=budget_ic,
            overrides=budget_overrides or None,
            cashbook=cashbook,
        )

        variance = (amount - compare_amount) if compare_amount is not None else None
        variance_pct = _variance_pct(amount, compare_amount) if compare_amount is not None else None
        prior_var = (amount - prior_amount) if prior_amount is not None else None
        py_var = (amount - py_amount) if py_amount is not None else None
        budget_var = (amount - budget_amount) if budget_amount is not None else None

        section_counters[layout.section] += 1
        prefix = section_refs.get(layout.section, "X")
        wp_ref = f"{prefix}.{section_counters[layout.section]}"
        from app.engines.working_papers import find_template

        acct_codes = [accounts[i].code for i in drill_ids if i in accounts]
        tmpl_codes = None if layout.line_code == CASH_XFER_CODE else acct_codes
        tmpl = find_template(line_code=layout.line_code, account_codes=tmpl_codes)
        if tmpl and not layout.is_total:
            wp_ref = tmpl.wp_ref
        elif tmpl and layout.line_code in ("NI", "NET_INCOME", "TOT_REV", "TOT_EXP", CURRENT_EARNINGS_CODE):
            wp_ref = tmpl.wp_ref
        drillable = bool(drill_ids)
        line_label = layout.line_label
        if cashbook and layout.line_code == EQUITY_CODE:
            line_label = "Opening equity"

        flag = _flag_flux(amount, prior_amount, mat_amt=mat_amt, mat_pct=mat_pct)
        lines.append(
            ReportLine(
                line_code=layout.line_code,
                line_label=line_label,
                section=layout.section,
                amount=amount,
                compare_amount=compare_amount,
                variance=variance,
                variance_pct=variance_pct,
                prior_period_amount=prior_amount,
                prior_period_variance=prior_var,
                prior_period_variance_pct=_variance_pct(amount, prior_amount),
                prior_year_amount=py_amount,
                prior_year_variance=py_var,
                prior_year_variance_pct=_variance_pct(amount, py_amount),
                budget_amount=budget_amount,
                budget_variance=budget_var,
                budget_variance_pct=_variance_pct(amount, budget_amount),
                flux_flag=flag,
                flux_note=_flux_note(line_label, amount, prior_amount, flag),
                indent_level=layout.indent_level,
                is_bold=layout.is_bold,
                is_total=layout.is_total,
                account_id=layout.account_id,
                drillable=drillable,
                account_ids=drill_ids,
                account_type_filter=type_filter,
                wp_ref=wp_ref if drillable else None,
            )
        )

    report = ReportOut(
        report_type=report_type,
        title=_statement_title(report_type),
        filters=filters,
        lines=lines,
        generated_at=datetime.utcnow().isoformat() + "Z",
        currency=filters.reporting_currency,
    )
    return _finalize_report(
        report,
        filters,
        prior_filters,
        py_filters,
        db=db,
        fx_missing=fx_missing,
        fx_missing_pairs=fx_missing_pairs,
    )


def _eval_formula(formula: str, line_amounts: dict[str, Decimal]) -> Decimal:
    """
    Safe subset: LINE_A + LINE_B - LINE_C
    """
    tokens = formula.replace("(", " ").replace(")", " ").replace("+", " + ").replace("-", " - ").split()
    if not tokens:
        return Decimal("0")
    total = Decimal("0")
    op = "+"
    for tok in tokens:
        if tok in ("+", "-"):
            op = tok
            continue
        val = line_amounts.get(tok, Decimal("0"))
        total = total + val if op == "+" else total - val
    return total


def _signed_compare(acct: DimAccount, totals: dict[int, Decimal] | None, report_type: str) -> Decimal | None:
    if totals is None:
        return None
    raw = totals.get(acct.id, Decimal("0"))
    return _signed_amount(acct, raw, report_type)


def _synthesize_report(
    db: Session,
    filters: ReportFilter,
    totals: dict[int, Decimal],
    compare_totals: dict[int, Decimal],
    accounts: dict[int, DimAccount],
    *,
    prior_totals: dict[int, Decimal] | None = None,
    py_totals: dict[int, Decimal] | None = None,
    budget_totals: dict[int, Decimal] | None = None,
    current_earnings: Decimal | None = None,
) -> ReportOut:
    type_order = {
        "income_statement": ["revenue", "expense"],
        "balance_sheet": ["asset", "liability", "equity"],
        "cash_flow": ["asset"],
    }
    wanted = type_order.get(filters.report_type, ["revenue", "expense", "asset", "liability", "equity"])
    lines: list[ReportLine] = []
    section_totals: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    prior_section: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    py_section: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    budget_section: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    compare_section: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    section_refs = {"revenue": "A", "expense": "B", "asset": "C", "liability": "D", "equity": "E", "totals": "Z"}
    section_counters: dict[str, int] = defaultdict(int)
    mat_amt, mat_pct = _materiality(filters)

    for acct_type in wanted:
        typed = [a for a in accounts.values() if a.account_type == acct_type and a.is_active]
        typed.sort(key=lambda a: (a.sort_order, a.code))
        typed_ids: list[int] = []
        for acct in typed:
            raw = totals.get(acct.id, Decimal("0"))
            amount = _signed_amount(acct, raw, filters.report_type)
            if amount == 0 and acct.id not in totals:
                continue
            compare_amount = (
                _signed_compare(acct, compare_totals, filters.report_type)
                if filters.compare_scenario_id
                else None
            )
            prior_amount = _signed_compare(acct, prior_totals, filters.report_type)
            py_amount = _signed_compare(acct, py_totals, filters.report_type)
            budget_amount = _signed_compare(acct, budget_totals, filters.report_type)
            section_totals[acct_type] += amount
            if prior_amount is not None:
                prior_section[acct_type] += prior_amount
            if py_amount is not None:
                py_section[acct_type] += py_amount
            if budget_amount is not None:
                budget_section[acct_type] += budget_amount
            if compare_amount is not None:
                compare_section[acct_type] += compare_amount
            typed_ids.append(acct.id)
            section_counters[acct_type] += 1
            from app.engines.working_papers import find_template

            tmpl = find_template(line_code=acct.code, account_codes=[acct.code])
            synth_ref = (
                tmpl.wp_ref
                if tmpl
                else f"{section_refs.get(acct_type, 'X')}.{section_counters[acct_type]}"
            )
            flag = _flag_flux(amount, prior_amount, mat_amt=mat_amt, mat_pct=mat_pct)
            lines.append(
                ReportLine(
                    line_code=acct.code,
                    line_label=acct.name,
                    section=acct_type,
                    amount=amount,
                    compare_amount=compare_amount,
                    variance=(amount - compare_amount) if compare_amount is not None else None,
                    prior_period_amount=prior_amount,
                    prior_period_variance=(amount - prior_amount) if prior_amount is not None else None,
                    prior_period_variance_pct=_variance_pct(amount, prior_amount),
                    prior_year_amount=py_amount,
                    prior_year_variance=(amount - py_amount) if py_amount is not None else None,
                    prior_year_variance_pct=_variance_pct(amount, py_amount),
                    budget_amount=budget_amount,
                    budget_variance=(amount - budget_amount) if budget_amount is not None else None,
                    budget_variance_pct=_variance_pct(amount, budget_amount),
                    flux_flag=flag,
                    flux_note=_flux_note(acct.name, amount, prior_amount, flag),
                    indent_level=1,
                    account_id=acct.id,
                    drillable=True,
                    account_ids=[acct.id],
                    wp_ref=synth_ref,
                )
            )
        if acct_type in section_totals:
            section_counters[acct_type] += 1
            tot = section_totals[acct_type]
            prior_t = prior_section.get(acct_type) if prior_totals is not None else None
            flag = _flag_flux(tot, prior_t, mat_amt=mat_amt, mat_pct=mat_pct)
            lines.append(
                ReportLine(
                    line_code=f"TOT_{acct_type.upper()}",
                    line_label=f"Total {acct_type.title()}",
                    section=acct_type,
                    amount=tot,
                    compare_amount=compare_section.get(acct_type) if filters.compare_scenario_id else None,
                    variance=(
                        tot - compare_section[acct_type] if filters.compare_scenario_id else None
                    ),
                    prior_period_amount=prior_t,
                    prior_period_variance=(tot - prior_t) if prior_t is not None else None,
                    prior_period_variance_pct=_variance_pct(tot, prior_t),
                    prior_year_amount=py_section.get(acct_type) if py_totals is not None else None,
                    budget_amount=budget_section.get(acct_type) if budget_totals is not None else None,
                    flux_flag=flag,
                    flux_note=_flux_note(f"Total {acct_type.title()}", tot, prior_t, flag),
                    is_bold=True,
                    is_total=True,
                    drillable=True,
                    account_ids=typed_ids,
                    account_type_filter=acct_type,
                    wp_ref=f"{section_refs.get(acct_type, 'X')}.{section_counters[acct_type]}",
                )
            )

    if filters.report_type == "income_statement":
        ni = section_totals.get("revenue", Decimal("0")) - section_totals.get("expense", Decimal("0"))
        ni_ids = [
            a.id for a in accounts.values() if a.account_type in ("revenue", "expense") and a.is_active
        ]
        prior_ni = None
        if prior_totals is not None:
            prior_ni = prior_section.get("revenue", Decimal("0")) - prior_section.get("expense", Decimal("0"))
        flag = _flag_flux(ni, prior_ni, mat_amt=mat_amt, mat_pct=mat_pct)
        lines.append(
            ReportLine(
                line_code="NET_INCOME",
                line_label="Net Income",
                section="totals",
                amount=ni,
                prior_period_amount=prior_ni,
                prior_period_variance=(ni - prior_ni) if prior_ni is not None else None,
                prior_period_variance_pct=_variance_pct(ni, prior_ni),
                flux_flag=flag,
                flux_note=_flux_note("Net Income", ni, prior_ni, flag),
                is_bold=True,
                is_total=True,
                drillable=True,
                account_ids=ni_ids,
                wp_ref="Z.1",
            )
        )

    if filters.report_type == "balance_sheet" and current_earnings is not None:
        eq = section_totals.get("equity", Decimal("0"))
        liab = section_totals.get("liability", Decimal("0"))
        tot_eq = eq + current_earnings
        lines.append(
            ReportLine(
                line_code=CURRENT_EARNINGS_CODE,
                line_label="Current earnings",
                section="equity",
                amount=current_earnings,
                indent_level=1,
                drillable=True,
                account_ids=_pnl_account_ids(accounts),
                wp_ref="Z.1",
            )
        )
        lines.append(
            ReportLine(
                line_code="BS_TOT_EQUITY",
                line_label="Total Equity",
                section="equity",
                amount=tot_eq,
                is_bold=True,
                is_total=True,
            )
        )
        lines.append(
            ReportLine(
                line_code="BS_TOT_L_AND_E",
                line_label="Total liabilities and equity",
                section="totals",
                amount=liab + tot_eq,
                is_bold=True,
                is_total=True,
            )
        )
        assets = section_totals.get("asset", Decimal("0"))
        lines.append(
            ReportLine(
                line_code="BS_TOT_ASSETS",
                line_label="Total Assets",
                section="asset",
                amount=assets,
                is_bold=True,
                is_total=True,
            )
        )

    return ReportOut(
        report_type=filters.report_type,
        title=_statement_title(filters.report_type),
        filters=filters,
        lines=lines,
        generated_at=datetime.utcnow().isoformat() + "Z",
        currency=filters.reporting_currency,
    )


def _finalize_report(
    report: ReportOut,
    filters: ReportFilter,
    prior_filters: ReportFilter | None,
    py_filters: ReportFilter | None,
    *,
    db: Session | None = None,
    fx_missing: bool = False,
    fx_missing_pairs: list[str] | None = None,
) -> ReportOut:
    columns = ["amount"]
    if filters.compare_prior_period:
        columns += ["prior_period", "prior_variance"]
    if filters.compare_prior_year:
        columns.append("prior_year")
    if filters.compare_budget:
        columns.append("budget")
    if filters.compare_scenario_id:
        columns += ["compare", "variance"]
    report.period_label = period_label(filters)
    report.prior_period_label = period_label(prior_filters) if prior_filters else None
    report.prior_year_label = period_label(py_filters) if py_filters else None
    report.budget_label = "Budget" if filters.compare_budget else None
    report.columns = columns
    report.fx_missing = fx_missing
    report.fx_missing_pairs = fx_missing_pairs or []
    entity_name = _entity_cover_name(db, filters) if db is not None else "Entity"
    report.entity_name = entity_name
    report.cover_title = (
        f"{entity_name} · {report.title} · {report.period_label} · {report.currency}"
    )
    if report.report_type == "balance_sheet":
        assets = _line_by_codes(report, ("BS_TOT_ASSETS", "TOT_ASSET"))
        le = _line_by_codes(report, ("BS_TOT_L_AND_E",))
        if assets is None:
            tot_a = next((l for l in report.lines if l.line_code == "TOT_ASSET"), None)
            assets = tot_a
        if assets is not None and le is not None:
            report.balance_difference = assets.amount - le.amount
            report.is_balanced = abs(report.balance_difference) < BALANCE_TOLERANCE
        elif assets is not None:
            liab = _line_by_codes(report, ("BS_TOT_LIAB", "TOT_LIABILITY"))
            eq = _line_by_codes(report, ("BS_TOT_EQUITY", "TOT_EQUITY"))
            if liab is not None and eq is not None:
                report.balance_difference = assets.amount - (liab.amount + eq.amount)
                report.is_balanced = abs(report.balance_difference) < BALANCE_TOLERANCE
    if not filters.include_zero_lines:
        report.lines = prune_zero_report_lines(report.lines)
    report.flux = _flux_items(report)
    return report


def _line_by_codes(report: ReportOut, codes: tuple[str, ...]) -> ReportLine | None:
    for line in report.lines:
        if line.line_code in codes:
            return line
    return None


def build_analytics_pack(db: Session, filters: ReportFilter) -> AnalyticsPack:
    """Analytical review: official P&L + BS + equity when one entity is in scope.

    Statement comparatives are prior fiscal year. Flux is month-over-month.
    Budget is omitted from the statutory pack (it is illustrative when seeded).
    """
    from .statement_pack import (
        PACK_DISCLAIMER,
        SCOPE_ERROR,
        assert_statement_scope,
        budget_is_illustrative,
        build_official_report,
        build_statement_diagnostics,
    )

    official = True
    scoped = filters
    try:
        eid = assert_statement_scope(db, filters)
        scoped = filters.model_copy(update={"entity_ids": [eid], "consolidate": False})
    except ValueError:
        official = False

    base = scoped.model_copy(
        update={
            "compare_prior_period": True,
            "compare_prior_year": True,
            "compare_budget": False,
        }
    )
    statements: list[ReportOut] = []
    if official:
        for rtype, period in (
            ("income_statement", filters.period or "ytd"),
            ("balance_sheet", "monthly"),
            ("equity", "ytd"),
        ):
            stmt_filters = base.model_copy(update={"report_type": rtype, "period": period})
            statements.append(build_official_report(db, stmt_filters))
        notes = list(statements[0].notes)
        disclaimer = PACK_DISCLAIMER
        try:
            can_print = build_statement_diagnostics(db, base).can_print
        except ValueError:
            can_print = False
    else:
        for rtype, period in (
            ("income_statement", filters.period or "ytd"),
            ("balance_sheet", "monthly"),
        ):
            stmt_filters = base.model_copy(update={"report_type": rtype, "period": period})
            statements.append(build_report(db, stmt_filters))
        notes = []
        disclaimer = SCOPE_ERROR + " " + PACK_DISCLAIMER
        can_print = False

    flux: list[FluxItem] = []
    for stmt in statements:
        flux.extend(stmt.flux)
    flux.sort(key=lambda i: abs(i.variance or Decimal("0")), reverse=True)

    mat_amt, mat_pct = _materiality(base)
    kpis: list[AnalyticsKpi] = []
    is_rpt = statements[0]
    bs_rpt = statements[1]
    for key, codes, label in (
        ("revenue", ("TOT_REV", "TOT_REVENUE"), "Revenue"),
        ("expense", ("TOT_EXP", "TOT_EXPENSE"), "Expenses"),
        ("net_income", ("NI", "NET_INCOME"), "Net income"),
        ("cash", ("BS_CASH", "CASH"), "Cash"),
    ):
        src = bs_rpt if key == "cash" else is_rpt
        line = _line_by_codes(src, codes)
        if not line:
            continue
        var = line.prior_period_variance
        tone = None
        if var is not None:
            favorable = var >= 0 if key in ("revenue", "net_income", "cash") else var <= 0
            tone = "ok" if favorable else "warn"
        kpis.append(
            AnalyticsKpi(
                key=key,
                label=label,
                amount=line.amount,
                prior_amount=line.prior_period_amount,
                variance=var,
                variance_pct=line.prior_period_variance_pct,
                tone=tone,
            )
        )

    return AnalyticsPack(
        period_label=period_label(base),
        currency=base.reporting_currency,
        materiality_amount=mat_amt,
        materiality_pct=mat_pct,
        kpis=kpis,
        flux=flux[:40],
        statements=statements,
        generated_at=datetime.utcnow().isoformat() + "Z",
        budget_is_illustrative=budget_is_illustrative(db),
        notes=notes,
        pack_disclaimer=disclaimer,
        can_print=can_print,
    )


def export_statement_pack_xlsx(db: Session, filters: ReportFilter) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from .statement_pack import PACK_DISCLAIMER, budget_is_illustrative, build_trial_balance

    pack = build_analytics_pack(db, filters)
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1D2E29")
    total_font = Font(bold=True)
    material_fill = PatternFill("solid", fgColor="3D2A12")
    thin = Border(
        left=Side(style="thin", color="7DA898"),
        right=Side(style="thin", color="7DA898"),
        top=Side(style="thin", color="7DA898"),
        bottom=Side(style="thin", color="7DA898"),
    )
    money_fmt = '#,##0.00;(#,##0.00);"—"'

    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Keystone Ledger"
    cover["A1"].font = Font(bold=True, size=16)
    cover["A2"] = "Unaudited financial reporting pack"
    cover["A3"] = pack.pack_disclaimer or PACK_DISCLAIMER
    cover["A5"] = "Period"
    cover["B5"] = pack.period_label
    cover["A6"] = "Currency"
    cover["B6"] = pack.currency
    cover["A7"] = "Printable"
    cover["B7"] = "Yes" if pack.can_print else "No — do not issue this pack"
    cover["A8"] = "Budget column"
    cover["B8"] = (
        "Illustrative target — not a statutory budget"
        if (pack.budget_is_illustrative or budget_is_illustrative(db))
        else "Posted budget lines"
    )
    cover["A10"] = "Notes"
    cover["A10"].font = total_font
    note_row = 11
    for note in pack.notes:
        cover.cell(note_row, 1, note.heading)
        cover.cell(note_row, 2, note.body)
        note_row += 1
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 80

    def write_sheet(ws, report: ReportOut) -> None:
        ws.append([report.cover_title or report.title])
        ws.append([f"{report.period_label or ''} · {report.currency}"])
        if report.report_type == "balance_sheet":
            if report.is_balanced:
                ws.append(["Balance check: in balance"])
            elif report.balance_difference is not None:
                ws.append([f"Balance check: out of balance by {float(report.balance_difference):,.2f}"])
            else:
                ws.append(["Balance check: n/a"])
        if report.fx_missing:
            ws.append([f"Missing FX rates: {', '.join(report.fx_missing_pairs)}"])
        ws.append([])
        headers = ["Ref", "Line", "Current"]
        if report.prior_period_label:
            headers += [f"Prior ({report.prior_period_label})", "$ Var", "% Var"]
        if report.prior_year_label:
            headers.append(f"Prior year ({report.prior_year_label})")
        if report.budget_label:
            headers += ["Budget", "vs Budget"]
        ws.append(headers)
        header_row = ws.max_row
        for col, _ in enumerate(headers, start=1):
            cell = ws.cell(header_row, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for line in report.lines:
            row = [line.wp_ref or "", line.line_label, float(line.amount)]
            if report.prior_period_label:
                row += [
                    float(line.prior_period_amount) if line.prior_period_amount is not None else None,
                    float(line.prior_period_variance) if line.prior_period_variance is not None else None,
                    float(line.prior_period_variance_pct) if line.prior_period_variance_pct is not None else None,
                ]
            if report.prior_year_label:
                row.append(float(line.prior_year_amount) if line.prior_year_amount is not None else None)
            if report.budget_label:
                row += [
                    float(line.budget_amount) if line.budget_amount is not None else None,
                    float(line.budget_variance) if line.budget_variance is not None else None,
                ]
            ws.append(row)
            r = ws.max_row
            if line.is_bold or line.is_total:
                for c in range(1, len(headers) + 1):
                    ws.cell(r, c).font = total_font
            if line.flux_flag:
                for c in range(1, len(headers) + 1):
                    ws.cell(r, c).fill = material_fill
            for c in range(3, len(headers) + 1):
                ws.cell(r, c).number_format = money_fmt
                ws.cell(r, c).border = thin
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22 if col > 1 else 10
        ws.column_dimensions["B"].width = 42
        ws.freeze_panes = f"C{header_row + 1}"

    for stmt in pack.statements:
        ws = wb.create_sheet(stmt.title[:31])
        write_sheet(ws, stmt)

    try:
        tb = build_trial_balance(db, filters)
        ws_tb = wb.create_sheet("Trial Balance")
        ws_tb.append([tb.cover_title or tb.title])
        ws_tb.append([tb.period_label or "", tb.accounting_basis or ""])
        ws_tb.append([])
        tb_headers = ["Code", "Account", "Type", "Debit", "Credit", "Mapped line", "Exception"]
        ws_tb.append(tb_headers)
        header_row = ws_tb.max_row
        for col, _ in enumerate(tb_headers, start=1):
            cell = ws_tb.cell(header_row, col)
            cell.font = header_font
            cell.fill = header_fill
        for row in tb.rows:
            ws_tb.append(
                [
                    row.account_code,
                    row.account_name,
                    row.account_type,
                    float(row.debit),
                    float(row.credit),
                    row.line_label or "",
                    row.exception or "",
                ]
            )
            r = ws_tb.max_row
            ws_tb.cell(r, 4).number_format = money_fmt
            ws_tb.cell(r, 5).number_format = money_fmt
        ws_tb.append(["", "Total", "", float(tb.total_debit), float(tb.total_credit), "", ""])
        tot_row = ws_tb.max_row
        for c in range(1, 8):
            ws_tb.cell(tot_row, c).font = total_font
        ws_tb.cell(tot_row, 4).number_format = money_fmt
        ws_tb.cell(tot_row, 5).number_format = money_fmt
        ws_tb.column_dimensions["B"].width = 42
        ws_tb.column_dimensions["F"].width = 28
        ws_tb.column_dimensions["G"].width = 40
    except ValueError:
        pass

    if pack.notes:
        notes_ws = wb.create_sheet("Notes")
        notes_ws["A1"] = "Notes to the unaudited pack"
        notes_ws["A1"].font = Font(bold=True, size=14)
        notes_ws["A2"] = pack.pack_disclaimer or PACK_DISCLAIMER
        notes_ws.append(["Heading", "Note"])
        for note in pack.notes:
            notes_ws.append([note.heading, note.body])
        notes_ws.column_dimensions["A"].width = 28
        notes_ws.column_dimensions["B"].width = 90

    flux_ws = wb.create_sheet("Flux")
    flux_ws.append(["Statement", "Ref", "Line", "Current", "Prior", "$ Var", "% Var", "Flag", "Commentary"])
    for col in range(1, 10):
        cell = flux_ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
    for item in pack.flux:
        flux_ws.append(
            [
                item.report_type,
                item.wp_ref or "",
                item.line_label,
                float(item.amount),
                float(item.prior_amount) if item.prior_amount is not None else None,
                float(item.variance) if item.variance is not None else None,
                float(item.variance_pct) if item.variance_pct is not None else None,
                item.flag,
                item.note,
            ]
        )
    flux_ws.column_dimensions["C"].width = 40
    flux_ws.column_dimensions["I"].width = 55

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
